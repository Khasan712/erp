import datetime
import os
import jwt
import xlwt
import openpyxl
from django.conf import settings
from django.http import HttpResponse
from django.core.exceptions import ValidationError
from django.contrib.sites.shortcuts import get_current_site
from django.core.files import File

from rest_framework import generics, status, permissions
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.response import Response

from django.db import transaction
from django.db.models import Avg, Sum, Count, Q, FloatField
from django.db.models.functions import Coalesce

from .tasks import send_message_to_suppliers
from api.v1.users.utils import send_message_register
from api.v1.users.models import User
from api.v1.users.services import make_errors
from api.v1.suppliers.models import Supplier
from api.v1.commons.pagination import make_pagination
from api.v1.commons.views import (
    string_to_date,
    exception_response,
    get_serializer_errors,
    raise_exception_response,
    object_not_found_response,
    serializer_valid_response,
    get_serializer_valid_response
)
from api.v1.sourcing.models import (
    CategoryRequest,
    SourcingCommentFile,
    SourcingComments,
    SourcingRequest,
    SourcingRequestEvent,
    SupplierAnswer,
    SupplierResult, SourcingRequestEventSuppliers, DocumentSourcing, SourcingRequestService
)
from api.v1.users.permissions import (
    IsSupplier,
    IsCategoryManager,
    IsSourcingDirector, IsContractAdministrator, IsSourcingAdministrator
)
from api.v1.chat.views import (
    send_result_notification
)

from .serializers import (
    CategoryRequestSerializer,
    EventCategorySerializer,
    EventDetailSerializer,
    EventInfoSerializer,
    EventQuestionarySerializer,
    EventSerializer,
    EventQuestionSerializer,
    SourcingCommentsSerializers,
    SourcingGetCommentsSerializers,
    SourcingRequestSerializer,
    SupplierSerializer,
    SupplierAnswerSerializers,
    CheckSASerializers,
    SupplierResultSerializer,
    GetEventByRequestIDSerializers,
    DocumentSourcingRequestSerializer,
    DocumentSourcingEventSerializer, SourcingRequestGetSerializer, SourcingRequestListSerializer,
    SourcingEventBySourcingRequestIDSerializers, SourcingEventSuppliersSerializer, SourcingEventInfosSerializer,
    SourcingEventQuestionarySerializer, SourcingRequestServiceSerializers, SourcingRequestCommoditySerializers,
    SourcingRequestConsultantSerializers, EventInfoUpdateSerializer, EventUpdateSerializer,
    EventSupplierUpdateSerializer, SourcingRequestAssignedListSerializer, SourcingEventSupplierQuestionarySerializer,
    SupplierAnswerInEventSerializer, SourcingCommentsQuestionarySerializer, SourcingCommentsQuestionaryGetSerializer
)


class CategoryRequestView(generics.ListCreateAPIView):
    permission_classes = (permissions.IsAuthenticated,)
    queryset = CategoryRequest.objects.select_related('organization')
    serializer_class = CategoryRequestSerializer

    def post(self, request):
        try:
            data = request.data
            user = self.request.user
            data['organization'] = user.organization.id
            serializers = self.get_serializer(data=data)
            if not serializers.is_valid():
                return Response(
                    {
                        "success": False,
                        "message": 'Error occured.',
                        "error": make_errors(serializers.errors),
                        "data": [],
                    }, status=status.HTTP_400_BAD_REQUEST
                )
            serializers.save()
            return Response(
                {
                    "success": True,
                    "message": 'User created suucessfully.',
                    "error": [],
                    "data": serializers.data.get('name'),
                }, status=status.HTTP_201_CREATED
            )
        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": 'Error occured.',
                    "error": str(e),
                    "data": [],
                },status=status.HTTP_400_BAD_REQUEST
            )

    def get(self, request):
        try:
            items = self.get_queryset().filter(organization_id=self.request.user.organization.pk)
            return Response(
                make_pagination(request, self.get_serializer(), self.get_queryset()),
                status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response(
                exception_response(e), status=status.HTTP_400_BAD_REQUEST
            )


class CategoryRequestListAPIView(generics.ListAPIView):
    permission_classes = (permissions.IsAuthenticated,)
    queryset = CategoryRequest.objects.select_related('organization')
    serializer_class = CategoryRequestSerializer

    def get(self, request):
        try:
            organization_id = self.request.user.organization.id
            items = self.get_queryset().filter(organization_id=organization_id)
            return Response(
                {
                    "success": True,
                    "message": 'Category got successfully.',
                    "error": [],
                    "data": self.get_serializer(items, many=True).data,
                }, status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response(
                exception_response(e), status=status.HTTP_400_BAD_REQUEST
            )


class SourcingRequestView(APIView):
    parser_classes = (FormParser, MultiPartParser, JSONParser)
    permission_classes = (
        permissions.IsAuthenticated, IsSourcingDirector | IsContractAdministrator | IsSourcingAdministrator |
        IsCategoryManager
    )

    def get_queryset(self):
        user = self.request.user
        organization_id = user.organization.id
        params = self.request.query_params
        items = SourcingRequest.objects.select_related(
            'organization', 'departement', 'categoryRequest', 'requestor', 'assigned_to', 'currency'
        ).filter(organization_id=organization_id)
        if params.get('q'):
            search = params.get('q')
            items = items.filter(
                Q(departement__name__icontains=search) | Q(categoryRequest__name__icontains=search) |
                Q(currency__name__icontains=search) | Q(estimated_budget__icontains=search) |
                Q(sourcing_request_name__icontains=search) | Q(sourcing_number__icontains=search) |
                Q(description__icontains=search) | Q(sourcing_request_status__icontains=search) |
                Q(serviceCommodityConsultant__icontains=search)
            )
        if params.get('request'):
            items = items.filter(id=params.get('request'))
        if params.get('currency'):
            items = items.filter(currency_id=params.get('currency'))
        if params.get('category'):
            items = items.filter(categoryRequest_id=params.get('category'))
        if params.get('from_budget'):
            items = items.filter(estimated_budget__gte=params.get('from_budget'))
        if params.get('to_budget'):
            items = items.filter(estimated_budget__lte=params.get('to_budget'))
        if params.get('created_date'):
            items = items.filter(created_at__date=string_to_date(params.get('created_date')).date())
        if params.get('deadline_date'):
            items = items.filter(deadline_at__date=string_to_date(params.get('deadline_date')).date())
        if params.get('sourcing_status'):
            items = items.filter(sourcing_request_status=params.get('sourcing_status'))
        return items

    def create_contract_assigned_users(self, assigned_users: list, sourcing_request_id: int):
        with transaction.atomic():
            for assigned_user in assigned_users:
                contract_assigned_user_serializer = SourcingRequestAssignedListSerializer(data=assigned_user)
                if not contract_assigned_user_serializer.is_valid():
                    get_serializer_errors(contract_assigned_user_serializer)
                contract_assigned_user_serializer.save(sourcingRequest_id=sourcing_request_id)

    def create_service_choices(self, service_choice, items, sourcing_request_id):
        match service_choice:
            case 'service':
                for service in items:
                    service['sourcingService'] = service.get('id')
                    service_serializer = SourcingRequestServiceSerializers(data=service)
                    if not service_serializer.is_valid():
                        get_serializer_errors(service_serializer)
                    service_serializer.save(sourcingRequest_id=sourcing_request_id, creator_id=self.request.user.id)
            case 'commodity':
                for commodity in items:
                    commodity['sourcingCommodity'] = commodity.get('id')
                    commodity_serializer = SourcingRequestCommoditySerializers(data=commodity)
                    if not commodity_serializer.is_valid():
                        get_serializer_errors(commodity_serializer)
                    commodity_serializer.save(sourcingRequest_id=sourcing_request_id, creator_id=self.request.user.id)
            case 'consultant':
                for consultant in items:
                    consultant['sourcingConsultant'] = consultant.get('id')
                    consultant_serializer = SourcingRequestConsultantSerializers(data=consultant)
                    if not consultant_serializer.is_valid():
                        get_serializer_errors(consultant_serializer)
                    consultant_serializer.save(sourcingRequest_id=sourcing_request_id, creator_id=self.request.user.id)

    def post(self, request):
        try:
            user = self.request.user
            data = request.data
            service_choice = data.get('serviceCommodityConsultant')
            items = data.get('items')
            assigned_users = data.get('assigned_to')
            data['requestor'] = user.id
            with transaction.atomic():
                serializer = SourcingRequestSerializer(data=data)
                if not serializer.is_valid():
                    raise ValidationError(message=f'{make_errors(serializer.errors)}')
                serializer.save(organization_id=user.organization.id)
                if assigned_users:
                    self.create_contract_assigned_users(assigned_users, serializer.data.get('id'))
                if data.get('documents'):
                    for document in data.get('documents'):
                        document['sourcingRequest'] = serializer.data.get('id')
                        doc_serializer = DocumentSourcingRequestSerializer(data=document)
                        if not doc_serializer.is_valid():
                            raise ValidationError(message=f'{make_errors(doc_serializer.errors)}')
                        doc_serializer.save()
                if not items and service_choice:
                    raise ValidationError(message="Please select any choices!")
                if service_choice and items:
                    self.create_service_choices(service_choice, items, serializer.data.get('id'))
            return Response(
                {
                    "success": True,
                    "message": 'Successfully created sourcing request.',
                    "error": [],
                    "data": [],
                }, status=status.HTTP_201_CREATED
            )
        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": 'Error occurred.',
                    "error": str(e),
                    "data": [],
                }, status=status.HTTP_400_BAD_REQUEST
            )

    def get(self, request):
        try:
            if not self.request.query_params.get('request'):
                serializer = SourcingRequestListSerializer
                return Response(
                    make_pagination(request, serializer, self.get_queryset()),
                    status=status.HTTP_200_OK
                )
            serializer = SourcingRequestGetSerializer
            return Response(
                make_pagination(request, serializer, self.get_queryset()),
                status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response(
                exception_response(e),
                status=status.HTTP_400_BAD_REQUEST
            )

    def patch(self, request):
        try:
            data = request.data
            params = request.query_params
            sourcing_request_id = params.get('request_id')
            sourcing_request = SourcingRequest.objects.get(id=sourcing_request_id)
            serializers = SourcingRequestSerializer(data=data, instance=sourcing_request, partial=True)
            if not serializers.is_valid():
                return Response(
                    {
                        "success": False,
                        "message": 'Error occurred.',
                        "error": make_errors(serializers.errors),
                        "data": [],
                    }, status=status.HTTP_400_BAD_REQUEST
                )
            serializers.save()
            return Response(
                {
                    "success": True,
                    "message": 'Successfully updated sourcing request.',
                    "error": [],
                    "data": serializers.data,
                }, status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": 'Error occurred.',
                    "error": str(e),
                    "data": [],
                }, status=status.HTTP_400_BAD_REQUEST
            )


class SourcingRequestEventView(APIView):
    permission_classes = (
        permissions.IsAuthenticated, IsSourcingDirector | IsContractAdministrator | IsSourcingAdministrator |
        IsCategoryManager
    )

    def get_queryset(self):
        sourcing_request_id = self.request.query_params.get('sourcing_request')
        items = SourcingRequestEvent.objects.select_related('sourcing_request', 'creator', 'parent').filter(
            sourcing_request__organization_id=self.request.user.organization.id)
        if sourcing_request_id:
            sourcing_request = SourcingRequest.objects.select_related(
                'organization', 'departement', 'categoryRequest', 'requestor', 'assigned_to', 'currency'
            ).filter(id=sourcing_request_id, organization_id=self.request.user.organization.id).first()
            events = items.filter(sourcing_request_id=sourcing_request.id)
            serializer = GetEventByRequestIDSerializers(events, many=True)
            return serializer
        serializer = GetEventByRequestIDSerializers(items, many=True)
        return serializer

    def post(self, request):
        try:
            data = request.data
            user = self.request.user
            sourcing_request = data.get('sourcing_request')
            sre_infos = data.get('sre_infos')
            suppliers = data.get('suppliers')
            documents = data.get('documents')
            sre_questionary = data.get('sre_questionary')
            if sre_questionary:
                sre_categories = sre_questionary.get('categories')
                if sre_categories:
                    if sum([int(category.get('weight')) for category in sre_categories]) != 100:
                        return Response(
                            {
                                "success": False,
                                "message": 'Error occurred.',
                                "error": "Categories weight's not equal with 100",
                                "data": [],
                            }, status=status.HTTP_400_BAD_REQUEST
                        )
                    for category in sre_categories:
                        if sum([int(question.get('weight')) for question in category.get('questions')]) != int(category.get('weight')):
                            return Response(
                                {
                                    "success": False,
                                    "message": 'Error occurred.',
                                    "error": "Questions weight's not equal with category weight's",
                                    "data": [],
                                }, status=status.HTTP_400_BAD_REQUEST
                            )

            with transaction.atomic():
                data['creator'] = user.id
                sourcing_event_serializers = EventSerializer(data=data)
                if not sourcing_event_serializers.is_valid():
                    raise ValidationError(message=f'{make_errors(sourcing_event_serializers.errors)}')
                sourcing_event_serializers.save()
                if suppliers:
                    for supplier in suppliers:
                        supplier['sourcingRequestEvent'] = sourcing_event_serializers.data.get('id')
                        supplier_serializers = SupplierSerializer(data=supplier)
                        if not supplier_serializers.is_valid():
                            raise ValidationError(make_errors(supplier_serializers.errors))
                        supplier_serializers.save()
                if sre_infos:
                    for sre_info in sre_infos:
                        sre_info['creator'] = user.id
                        sre_info['parent'] = sourcing_event_serializers.data.get('id')
                        sre_info_serializers = EventInfoSerializer(data=sre_info)
                        if not sre_info_serializers.is_valid():
                            raise ValidationError(make_errors(sre_info_serializers.errors))
                        sre_info_serializers.save(sourcing_request_id=sourcing_request)
                if documents:
                    for document in documents:
                        doc_serializers = DocumentSourcingEventSerializer(data=document)
                        if not doc_serializers.is_valid():
                            raise ValidationError(make_errors(doc_serializers.errors))
                        doc_serializers.save(sourcingEvent_id=sourcing_event_serializers.data.get('id'))
                if sre_questionary:
                    sre_questionary['creator'] = user.id
                    sre_questionary['parent'] = sourcing_event_serializers.data.get('id')
                    sre_questionary_serializers = EventQuestionarySerializer(data=sre_questionary)
                    if not sre_questionary_serializers.is_valid():
                        raise ValidationError(make_errors(sre_questionary_serializers.errors))
                    sre_questionary_serializers.save(sourcing_request_id=sourcing_request)
                    if sre_questionary.get('categories'):
                        for sre_category in sre_questionary.get('categories'):
                            sre_category['creator'] = user.id
                            sre_category['parent'] = sre_questionary_serializers.data.get('id')
                            sre_category_serializers = EventCategorySerializer(data=sre_category)
                            if not sre_category_serializers.is_valid():
                                raise ValidationError(make_errors(sre_category_serializers.errors))
                            sre_category_serializers.save(sourcing_request_id=sourcing_request)
                            if sre_category.get('questions'):
                                for sre_question in sre_category.get('questions'):
                                    sre_question['creator'] = user.id
                                    sre_question['parent'] = sre_category_serializers.data.get('id')
                                    sre_question_serializers = EventQuestionSerializer(data=sre_question)
                                    if not sre_question_serializers.is_valid():
                                        raise ValidationError(make_errors(sre_question_serializers.errors))
                                    sre_question_serializers.save(sourcing_request_id=sourcing_request)
                current_site = get_current_site(request).domain
                send_message_to_suppliers(
                    current_site, sourcing_event_serializers.data.get('id')
                )
            return Response(
                {
                    "success": True,
                    "message": 'Sourcing requests event created successfully.',
                    "error": [],
                    "data": sourcing_event_serializers.data,
                }, status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": 'Error occurred.',
                    "error": str(e),
                    "data": [],
                }, status=status.HTTP_400_BAD_REQUEST
            )

    def get(self, request):
        try:
            return Response(
                {
                    "success": True,
                    "message": 'Sourcing request events.',
                    "error": [],
                    "data": self.get_queryset().data,
                }, status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": 'Error occurred.',
                    "error": str(e),
                    "data": [],
                }, status=status.HTTP_400_BAD_REQUEST
            )


class SourcingEventGetByParamsAPIView(APIView):
    permission_classes = (
        permissions.IsAuthenticated, IsSourcingDirector | IsContractAdministrator | IsSupplier |
        IsSourcingAdministrator | IsCategoryManager
    )

    def get_queryset(self):
        sourcing_events = SourcingRequestEvent.objects.select_related('sourcing_request', 'creator', 'parent').filter(
            sourcing_request__organization_id=self.request.user.organization.id
        )
        return sourcing_events

    def get_filter(self):
        user = self.request.user
        queryset = self.get_queryset()
        params = self.request.query_params
        sourcing_request = params.get('sourcing-request')
        event_id = params.get('event')
        suppliers_info_questionnaries = params.get('get_data')
        if sourcing_request:
            sourcing_events = queryset.filter(sourcing_request_id=sourcing_request,
                                                     general_status='sourcing_event')
            return sourcing_events
        match suppliers_info_questionnaries:
            case 'supplier':
                supplier_event = SourcingRequestEventSuppliers.objects.select_related(
                    'supplier', 'sourcingRequestEvent'
                ).filter(sourcingRequestEvent_id=event_id, supplier__parent_supplier_id=user.id).first()
                return supplier_event
            case 'suppliers':
                suppliers = SourcingRequestEventSuppliers.objects.select_related(
                    'supplier', 'sourcingRequestEvent'
                ).filter(sourcingRequestEvent_id=event_id)
                return suppliers
            case 'infos':
                infos = queryset.filter(parent_id=event_id, general_status='info')
                return infos
            case 'questionaries':
                questionaries = queryset.filter(parent_id=event_id, general_status='questionary').first()
                return questionaries
            case 'documents':
                documents = DocumentSourcing.objects.select_related("sourcingRequest", "sourcingEvent").filter(
                    sourcingEvent_id=event_id
                )
                return documents

    def get_supplier_answer_question(self, questionnaire: int, supplier: int):
        supplier_answers = SupplierAnswer.objects.select_related('supplier', 'question', 'supplier_result', 'checker').filter(
            supplier_id=supplier
        )
        categories = []
        for category in self.get_queryset().filter(parent_id=questionnaire, general_status='category'):
            category_questions = []
            for question in self.get_queryset().filter(parent_id=category.id, general_status='question'):
                supplier_answer = supplier_answers.filter(question__general_status='question', question_id=question.id).first()
                answered = None
                if supplier_answer:
                    answered = {
                        'id': supplier_answer.question.id,
                        'answer_type': 'yes_no' if supplier_answer.yes_no is not None else 'input',
                        'answer': supplier_answer.yes_no if supplier_answer.yes_no is not None else supplier_answer.answer,
                    }
                question_obj = {
                    'id': question.id,
                    'text': question.text,
                    'weight': question.weight,
                    'answer_type': 'yes_no' if question.yes_no is not None else 'input',
                    'answered': answered
                }
                category_questions.append(question_obj)
            category_obj = {
                'id': category.id,
                'title': category.title,
                'weight': category.weight,
                'questions': category_questions
            }
            categories.append(category_obj)
        return categories

    def get_serializer(self):
        user = self.request.user
        filtered_data = self.get_filter()
        params = self.request.query_params
        event_id = params.get('event')
        sourcing_request = params.get('sourcing-request')
        suppliers_info_questionnaires = params.get('get_data')
        if sourcing_request:
            return SourcingEventBySourcingRequestIDSerializers(filtered_data, many=True).data
        match suppliers_info_questionnaires:
            case 'supplier':
                return {
                    'id': filtered_data.supplier.id,
                    'name': filtered_data.supplier.name
                }
            case 'suppliers':
                return SourcingEventSuppliersSerializer(filtered_data, many=True).data
            case 'infos':
                return SourcingEventInfosSerializer(filtered_data, many=True).data
            case 'questionaries':
                if user.role == 'supplier':
                    supplier_event = SourcingRequestEventSuppliers.objects.select_related(
                        'supplier', 'sourcingRequestEvent'
                    ).filter(sourcingRequestEvent_id=event_id, supplier__parent_supplier_id=user.id).first()
                    if supplier_event is not None and supplier_event.supplier_timeline == 'not_viewed':
                        supplier_event.supplier_timeline = 'in_progress'
                        supplier_event.save()
                    supplier_answer_questions = self.get_supplier_answer_question(filtered_data.id, supplier_event.supplier.id)
                    serializer = SourcingEventSupplierQuestionarySerializer(filtered_data).data
                    questionnaire = serializer.copy()
                    questionnaire['get_questionary_data'] = {
                        'answers_of_question': supplier_answer_questions,
                        'timeline': supplier_event.supplier_timeline,
                        'is_submitted': True if supplier_event.supplier_timeline in ['done', 'passed', 'rejected'] else False
                    }
                    return questionnaire
                return SourcingEventQuestionarySerializer(filtered_data).data
            case 'documents':
                return SourcingEventQuestionarySerializer(filtered_data).data

    def get(self, request):
        try:
            return Response(
                {
                    "success": True,
                    "message": 'Successfully got data.',
                    "error": [],
                    "data": self.get_serializer(),
                }, status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": 'Error occurred.',
                    "error": str(e),
                    "data": [],
                }, status=status.HTTP_400_BAD_REQUEST
            )


class SourcingRequestEventDetailView(APIView):
    permission_classes = (permissions.IsAuthenticated, IsSourcingDirector | IsSourcingAdministrator)

    def get_queryset(self):
        queryset = SourcingRequestEvent.objects.select_related('sourcing_request', 'creator', 'parent').filter(
            sourcing_request__organization_id=self.request.user.organization.id
        )
        return queryset

    def get_object(self, pk):
        return self.get_queryset().filter(id=pk).first()

    def update_infos(self, infos, event):
        sre_infos = []
        for sre_info in infos:
            info = self.get_queryset().filter(id=sre_info['id'], parent_id=event).first()
            if info is not None:
                sre_info_serializers = EventInfoUpdateSerializer(info, data=sre_info, partial=True)
                if not sre_info_serializers.is_valid():
                    get_serializer_errors(sre_info_serializers)
                sre_info_serializers.save()
            else:
                sre_info_serializers = EventInfoUpdateSerializer(data=sre_info)
                if not sre_info_serializers.is_valid():
                    get_serializer_errors(sre_info_serializers)
                sre_info_serializers.save(
                    creator_id=self.request.user.id,
                    parent_id=event, sourcing_request_id=self.get_object(event).sourcing_request.id)
            sre_infos.append(sre_info_serializers.data)
        return sre_infos

    def update_suppliers(self, suppliers, sourcing_event):
        event_suppliers = []
        for supplier in suppliers:
            event_supplier = SourcingRequestEventSuppliers.objects.select_related(
                'supplier', 'sourcingRequestEvent').filter(sourcingRequestEvent_id=sourcing_event,
                                                           supplier_id=supplier['id']).first()
            supplier['supplier'] = supplier['id']
            if event_supplier is not None:
                supplier_serializers = EventSupplierUpdateSerializer(event_supplier, data=supplier, partial=True)
                if not supplier_serializers.is_valid():
                    get_serializer_errors(supplier_serializers)
                supplier_serializers.save()
            else:
                supplier_serializers = EventSupplierUpdateSerializer(data=supplier)
                if not supplier_serializers.is_valid():
                    raise ValidationError(make_errors(supplier_serializers.errors))
                supplier_serializers.save(sourcingRequestEvent_id=sourcing_event)
            event_suppliers.append(supplier_serializers.data)
        return event_suppliers

    def update_questions(self, questions: list, category: dict, event: int):
        if sum([int(question.get('weight')) for question in questions]) != category.get('weight'):
            raise ValidationError(message='Questions weight\'s not equal with category weight!')
        category_questions = []
        with transaction.atomic():
            for question in questions:
                question_id = self.get_queryset().filter(id=question.get('id'), parent_id=category.get('id')).first()
                if question_id is not None:
                    question_serializer = EventQuestionSerializer(question_id, data=question, partial=True)
                    if not question_serializer.is_valid():
                        get_serializer_errors(question_serializer)
                    question_serializer.save()
                else:
                    question['creator'] = self.request.user.id
                    question['parent'] = category.get('id')
                    sourcing_event = self.get_object(event)
                    question_serializer = EventQuestionSerializer(data=question)
                    if not question_serializer.is_valid():
                        get_serializer_errors(question_serializer)
                    question_serializer.save(sourcing_request_id=sourcing_event.sourcing_request.id)
                category_questions.append(question_serializer.data)
        return category_questions

    def update_category(self, categories: list, questionary: int, event: int):
        if sum([int(category.get('weight')) for category in categories]) != 100:
            raise ValidationError(message='Categories weight\'s more then 100!')
        questionnaire_categories = []
        with transaction.atomic():
            for category in categories:
                category_id = self.get_queryset().filter(id=category.get('id'), parent_id=questionary).first()
                if category_id is not None:
                    category_serializer = EventCategorySerializer(category_id, data=category, partial=True)
                    if not category_serializer.is_valid():
                        get_serializer_errors(category_serializer)
                    category_serializer.save()
                    questionnaire_category = category_serializer.data.copy()
                else:
                    category['creator'] = self.request.user.id
                    category['parent'] = questionary
                    sourcing_event = self.get_object(event)
                    category_serializer = EventCategorySerializer(data=category)
                    if not category_serializer.is_valid():
                        get_serializer_errors(category_serializer)
                    category_serializer.save(sourcing_request_id=sourcing_event.sourcing_request.id)
                    questionnaire_category = category_serializer.data.copy()
                category_questions = self.update_questions(category.get('questions'), category_serializer.data, event)
                questionnaire_category['questions'] = category_questions
                questionnaire_categories.append(questionnaire_category)
        return questionnaire_categories

    def update_questionnaire(self, questionnaire, event):
        if questionnaire.get('success_weight') > 100:
            raise ValidationError(message='Success weight more then 100.')
        questionnaire_id = self.get_queryset().get(id=questionnaire.get('id'), parent_id=event)
        questionnaire_serializer = EventQuestionarySerializer(questionnaire_id, data=questionnaire, partial=True)
        if not questionnaire_serializer.is_valid():
            get_serializer_errors(questionnaire_serializer)
        questionnaire_serializer.save()
        event_questionnaire = questionnaire_serializer.data.copy()
        categories = questionnaire.get('categories')
        if categories:
            questionnaire_categories = self.update_category(categories, questionnaire.get('id'), event)
            event_questionnaire['categories'] = questionnaire_categories
        return event_questionnaire

    def get(self, request, id):
        try:
            token = request.GET.get('token')
            payload = jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'])
            user = User.objects.get(id=payload['user_id'])
            sourcing_event = SourcingRequestEvent.objects.get(id=id)
            if user.organization.id != sourcing_event.sourcing_request.organization.id:
                return Response(
                    {
                        "success": False,
                        "message": 'Error occurred.',
                        "error": "Be niece human...",
                        "data": [],
                    }, status=status.HTTP_400_BAD_REQUEST
                )
            serializer = EventDetailSerializer(sourcing_event)
            return Response(
                {
                    "success": True,
                    "message": 'Sourcing request events.',
                    "error": [],
                    "data": serializer.data,
                }, status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": 'Error occurred.',
                    "error": str(e),
                    "data": [],
                }, status=status.HTTP_400_BAD_REQUEST
            )

    def patch(self, request, id):
        try:
            sourcing_event = self.get_object(id)
            if not sourcing_event:
                return Response(object_not_found_response())
            data = self.request.data
            with transaction.atomic():
                sourcing_event_serializer = EventUpdateSerializer(sourcing_event, data=data, partial=True)
                if not sourcing_event_serializer.is_valid():
                    get_serializer_errors(sourcing_event_serializer)
                sourcing_event_serializer.save()
                infos = data.get('sre_infos')
                questionnaires = data.get('sre_questionary')
                suppliers = data.get('suppliers')
                sourcing_event = sourcing_event_serializer.data.copy()
                if infos is not None:
                    event_infos = self.update_infos(infos, id)
                    sourcing_event['infos'] = event_infos
                if questionnaires is not None:
                    event_questionnaire = self.update_questionnaire(questionnaires, id)
                    sourcing_event['questionnaire'] = event_questionnaire
                if suppliers is not None:
                    event_suppliers = self.update_suppliers(suppliers, id)
                    sourcing_event['suppliers'] = event_suppliers
        except Exception as e:
            return Response(
                exception_response(e), status=status.HTTP_400_BAD_REQUEST
            )
        return Response(
            {
                'success': True,
                'error': [],
                'message': "Successfully updated",
                'data': sourcing_event
            }, status=status.HTTP_200_OK
        )


class SourcingCommentsView(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def get_event_queryset(self):
        return SourcingRequestEvent.objects.select_related('sourcing_request', 'creator', 'parent')

    def get_request_queryset(self):
        return SourcingRequest.objects.select_related('organization', 'departement', 'categoryRequest', 'requestor', 'assigned_to', 'currency')

    def get_event_suppliers_queryset(self):
        return SourcingRequestEventSuppliers.objects.select_related('supplier', 'sourcingRequestEvent')

    def get_comments_queryset(self):
        return SourcingComments.objects.select_related('sourcingRequest', 'sourcingRequestEvent', 'author', 'supplier')

    def set_comment_files(self, files, comment_id, user_id):
        comment_files = []
        for file in files:
            f_time = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            file_name = f"{f_time}_{file.name}"
            saved_file_path = os.path.join('media', 'sourcing', 'comment', 'files', file_name)
            os.makedirs(os.path.dirname(saved_file_path), exist_ok=True)
            with open(saved_file_path, 'wb') as destination:
                for chunk in file.chunks():
                    destination.write(chunk)
            comment_files.append(
                SourcingCommentFile(
                    comment_id=comment_id,
                    creator_id=user_id,
                    uploaded_file=saved_file_path
                )
            )
        SourcingCommentFile.objects.bulk_create(comment_files)

    def post(self, request):
        try:
            user = self.request.user
            data = request.data
            params = self.request.query_params
            method = params.get("method")
            files = request.FILES.getlist('file')
            match method:
                case 'questionary':
                    if user.role == 'supplier':
                        questionary = self.get_event_queryset().filter(id=data.get('questionary'), general_status='questionary').first()
                        event_supplier = self.get_event_suppliers_queryset().filter(
                            sourcingRequestEvent_id=questionary.parent.id, supplier__supplier_id=user.id
                        ).first()
                        with transaction.atomic():
                            serializer = SourcingCommentsQuestionarySerializer(data=data)
                            if not serializer.is_valid():
                                raise ValueError(make_errors(serializer.errors))                            
                            serializer.save(
                                sourcingRequestEvent=questionary,
                                supplier=event_supplier.supplier,
                                author=user
                            )
                            if files:
                                self.set_comment_files(files, serializer.data.get('id'), user.id)
                        return Response({
                            'success': True
                        }, status=status.HTTP_201_CREATED)

                    questionary = self.get_event_queryset().filter(id=data.get('questionary')).first()
                    serializer = SourcingCommentsQuestionarySerializer(data=data)
                    if not serializer.is_valid():
                        raise ValueError(make_errors(serializer.errors))
                    serializer.save(
                        sourcingRequestEvent=questionary,
                        supplier_id=data.get("supplier"),
                        author=user
                    )
                    if files:
                        self.set_comment_files(files, serializer.data.get('id'), user.id)
                    return Response({
                        'success': True
                    }, status=status.HTTP_201_CREATED)
                case 'sourcing.request':
                    sourcing_request = self.get_request_queryset().filter(
                        id=data.get('request'), organization_id=user.organization.id
                    ).first()
                    serializer = SourcingCommentsQuestionarySerializer(data=data)
                    if not serializer.is_valid():
                        raise ValueError(make_errors(serializer.errors))
                    serializer.save(
                        sourcingRequest=sourcing_request,
                        author=user
                    )
                    return Response({
                        'success': True
                    }, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": 'Error occurred.',
                    "error": str(e),
                    "data": [],
                }, status=status.HTTP_400_BAD_REQUEST
            )
    def get(self, request):
        try:
            params = request.query_params
            method = params.get("method")
            user = self.request.user
            match method:
                case 'sourcing.request':
                    s_request_c = SourcingComments.objects.filter(
                        sourcingRequest_id=params.get('request'), sourcingRequest__organization_id=user.organization.id
                    )
                    serializers = SourcingGetCommentsSerializers(s_request_c, many=True)
                    return Response(
                        {
                            "success": True,
                            "data": serializers.data,
                        }, status=status.HTTP_200_OK
                    )
                case 'questionary':
                    if user.role == 'supplier':
                        questionary_comments = self.get_comments_queryset().filter(
                            sourcingRequestEvent_id=params.get('questionary'), supplier__supplier_id=user.id
                        ).order_by("-id")
                        serializer = SourcingCommentsQuestionaryGetSerializer
                        return Response({
                            "success": True,
                            'data': make_pagination(request, serializer, questionary_comments)
                        })

                    questionary_comments = self.get_comments_queryset().filter(
                        sourcingRequestEvent_id=params.get('questionary'), supplier_id=params.get('supplier')
                    ).order_by("-id")
                    serializer = SourcingCommentsQuestionaryGetSerializer
                    return Response({
                        "success": True,
                        'data': make_pagination(request, serializer, questionary_comments)
                    })
        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": 'Error occurred.',
                    "error": str(e),
                    "data": [],
                }, status=status.HTTP_400_BAD_REQUEST
            )


class SupplierAnswerView(APIView):
    permission_classes = (permissions.IsAuthenticated, )

    def give_weight(self):
        user = self.request.user
        data = self.request.data
        supplier_id = data.get('supplier_id')
        event_id = data.get('event_id')
        is_submitted = data.get('is_submitted')
        answers = data.get('answers')

        event = SourcingRequestEvent.objects.select_related('sourcing_request', 'creator', 'parent').filter(
            id=event_id, creator__organization_id=user.organization.id, general_status='sourcing_event'
        ).first()
        if not event:
            raise ValidationError('Sourcing event not found!')

        supplier_in_event = SourcingRequestEventSuppliers.objects.select_related('supplier', 'sourcingRequestEvent').filter(
            supplier_id=supplier_id, sourcingRequestEvent_id=event_id
        ).first()
        if not supplier_in_event:
            raise ValidationError("Supplier not found!")

        supplier_answers = SupplierAnswer.objects.select_related('supplier', 'question', 'supplier_result', 'checker').filter(
            supplier_id=supplier_id, question__parent__parent__parent_id=event_id
        )
        if not supplier_answers.exists():
            raise ValidationError("Supplier answers not found!")
        questionary = SourcingRequestEvent.objects.select_related('sourcing_request', 'creator', 'parent').filter(
            parent_id=event_id, creator__organization_id=user.organization.id, general_status='questionary'
        ).first()
        if not questionary:
            raise ValidationError("Questionary not found!")
        supplier_result, _ = SupplierResult.objects.select_related('questionary', 'supplier').get_or_create(
            questionary_id=questionary.id,
            supplier_id=supplier_in_event.supplier.id
        )
        if supplier_result.is_submitted:
            raise ValidationError("Supplier answers has already checked!")
        else:
            with transaction.atomic():
                for supplier_answer in supplier_answers.filter(question__yes_no__in=[True, False]):
                    if supplier_answer.yes_no != supplier_answer.question.yes_no:
                        supplier_answer.weight = supplier_answer.question.weight
                        supplier_answer.yes_no = supplier_answer.question.yes_no
                        supplier_answer.save()
                if answers:
                    for answer in answers:
                        supplier_answer = supplier_answers.filter(
                            question_id=answer['question_id'], question__yes_no=None
                        ).first()
                        if not supplier_answer:
                            raise ValidationError(message="Supplier answer not found.")
                        if float(answer['weight']) > supplier_answer.question.weight:
                            raise ValidationError(message="Weight is grater then question weight")
                        if float(answer['weight']) != supplier_answer.weight:
                            supplier_answer.weight = float(answer['weight'])
                            supplier_answer.save()

        if is_submitted:
            total_weight = supplier_answers.aggregate(foo=Coalesce(Sum('weight', output_field=FloatField()), 0.0))['foo']
            supplier_result.is_submitted = True
            supplier_result.total_weight = total_weight
            supplier_result.questionary_status = 'congratulations' if total_weight >= supplier_result.questionary.success_weight else 'rejected'
            supplier_result.save()
            # send_result_notification(supplier_result, user.id)

    def post(self, request):
        try:
            data = request.data
            if data.get('method') == 'give.weight':
                self.give_weight()
                return Response(
                    {
                        "success": True,
                        "message": 'Successfully checked answers.',
                        "error": [],
                        "data": [],
                    }, status=status.HTTP_200_OK
                )
            with transaction.atomic():
                supplier = Supplier.objects.select_related('organization', 'create_by', 'supplier', 'parent', 'supplier_result').filter(
                    id=data['supplier']
                ).first()
                supplier_in_event = SourcingRequestEventSuppliers.objects.select_related(
                    'supplier', 'sourcingRequestEvent'
                ).get(supplier_id=supplier.id, sourcingRequestEvent_id=data['event_id'])
                if supplier_in_event.supplier_timeline in ['done', 'passed', 'rejected']:
                    raise ValidationError(message='You have already submitted.')
                for d in data['answers']:
                    supplier_answer, _ = SupplierAnswer.objects.get_or_create(question_id=d['question'], supplier_id=data['supplier'])
                    if supplier_answer is not None:
                        supplier_answer_serializer = SupplierAnswerSerializers(supplier_answer, data=d, partial=True)
                        if not supplier_answer_serializer.is_valid():
                            raise ValidationError(message=f'{make_errors(supplier_answer_serializer.errors)}')
                        supplier_answer_serializer.save()
                    else:
                        d['supplier'] = supplier.id
                        serializer = SupplierAnswerSerializers(data=d)
                        if not serializer.is_valid():
                            raise ValidationError(message=f"{make_errors(serializer.errors)}")
                        serializer.save()
                if data['is_submitted']:
                    supplier_in_event.supplier_timeline = 'done'
                    supplier_in_event.save()
        except Exception as e:
            return Response(exception_response(e), status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response(
                {
                    "success": True,
                    "message": 'Successfully created answers.',
                    "error": [],
                    "data": [],
                }, status=status.HTTP_201_CREATED
            )

    def get(self, request):
        try:
            user = self.request.user
            supplier_id = self.request.data.get('supplier')
            supplier = Supplier.objects.select_related('organization', 'create_by', 'supplier', 'parent').filter(
                id=supplier_id, organization_id=user.organization.id,
            ).first()
            questionary_id = self.request.query_params.get('questionary')
            questionary = SourcingRequestEvent.objects.select_related('sourcing_request', 'creator', 'parent').filter(
                id=questionary_id, general_status='questionary'
            ).first()
            supplier_answer = SupplierAnswer.objects.select_related('supplier', 'question', 'supplier_result', 'checker').filter(
                supplier_id=supplier.id, question__parent__parent_id=questionary.id
            )
            # data = questionary.annotate(categories=)
        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": 'Error occurred.',
                    "error": str(e),
                    "data": [],
                }, status=status.HTTP_400_BAD_REQUEST
            )


class CheckSupplierAnswers(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def post(self, request, *args, **kwargs):
        try:
            pass
        except Exception as e:
            return Response(exception_response(e), status=status.HTTP_400_BAD_REQUEST)

class MassUpload(APIView):
    permission_classes = (permissions.IsAuthenticated,)
    def post(self, request):
        data = request.data
        filehandle = data.get("file")
        workbook = openpyxl.load_workbook(filehandle)
        worksheet = workbook["Sheet1"]
        excel_data = list()
        for i, row in enumerate(worksheet.iter_rows()):
            if i > 0:
                row_data = dict()
                inx = 0
                for cell in row:
                    n = list(worksheet.iter_rows())[0][inx]
                    row_data[n.value] = str(cell.value)
                    inx+=1
                # inx = 0
                excel_data.append(row_data)
        with transaction.atomic():
            for ex in excel_data:
                user = User(
                    email=ex.get("email"),
                    phone=ex.get("phone"),
                    organization_id=ex.get("organization"),
                    first_name=ex.get("first_name"),
                    last_name=ex.get("last_name"),
                )
                user.set_password("1")
                user.save()
                user_data = {
                    'first_name': user.first_name,
                    'email': user.email,
                    'role': user.role,
                    'organization': self.request.user.organization.id,
                    'organization_name': self.request.user.organization.name,
                    'password': ex.get("password"),
                }
                send_message_register(request, user_data)
        return Response(User.objects.values('email'))


def convert_to_excel(users):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="users.csv"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Users')

    # Sheet header, first row
    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True
    columns = ('id', 'first_name', 'last_name', 'email',)
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    # users = User.objects.values_list('id', 'first_name', 'last_name', 'email')
    for row in users:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)

    wb.save(response)
    return response
class MassDownload(APIView):
    def post(self, request):
        params = self.request.query_params

        users = User.objects.filter(
            organization_id=self.request.user.organization.id
        ).values_list('id', 'first_name', 'last_name', 'email')
        return convert_to_excel(users)


class SourcingRequestStatusStatisticsView(APIView):
    permission_classes = (permissions.IsAuthenticated, IsSourcingDirector | IsSourcingAdministrator)

    def get(self, request):
        try:
            user = self.request.user
            request_status = SourcingRequest.objects.select_related(
                'organization', 'departement', 'categoryRequest', 'requestor', 'assigned_to', 'currency'
            ).filter(organization_id=user.organization.id)
            data = {
                "total": request_status.count(),
                "created": request_status.filter(sourcing_request_status='created').count(),
                "assigned": request_status.filter(sourcing_request_status='assigned').count(),
                "in_progress": request_status.filter(sourcing_request_status='in_progress').count(),
                "outdated": request_status.filter(sourcing_request_status='outdated').count(),
                "completed": request_status.filter(sourcing_request_status='completed').count(),
                "closed": request_status.filter(sourcing_request_status='closed').count(),
            }
        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": 'Error occurred.',
                    "error": str(e),
                    "data": [],
                }, status=status.HTTP_400_BAD_REQUEST
            )
        return Response(
            {
                "success": True,
                "message": 'Successfully got sourcing requests statistic\'s.',
                "error": [],
                "data": data,
            }, status=status.HTTP_200_OK
        )


def get_sourcing_request_queryset():
    return SourcingRequest.objects.select_related(
        'organization', 'departement', 'categoryRequest', 'requestor', 'assigned_to', 'currency'
    )


def sourcing_request_list_category_manager(request, category_manager):
    sourcing_request = SourcingRequest.objects.select_related(
            'organization', 'departement', 'categoryRequest', 'requestor', 'assigned_to', 'currency'
        ).filter(
        Q(requestor_id=category_manager.id) | Q(assigned_to_id=category_manager.id) |
        Q(assigned_sourcing_request__assigned_id=category_manager.id)
    )
    return make_pagination(request, SourcingRequestListSerializer, sourcing_request)


def get_supplier_answers(request):
    sourcing_request = request.data.get('parametr')
    supplier_id = sourcing_request.get('supplier_id')
    event_id = sourcing_request.get('event_id')

    # Queryset
    event_queryset = SourcingRequestEvent.objects.select_related('sourcing_request', 'creator', 'parent')
    supplier = Supplier.objects.select_related('organization', 'create_by', 'supplier', 'parent_supplier').filter(
        id=supplier_id
    ).first()
    if not supplier:
        raise ValidationError("Supplier not found!")
    questionary = event_queryset.filter(parent_id=event_id, general_status='questionary').first()
    if not questionary:
        raise ValidationError('Questionary not found!')
    supplier_result, _ = SupplierResult.objects.select_related('questionary', 'supplier').get_or_create(
        questionary_id=questionary.id,
        supplier_id=supplier_id
    )
    categories = []
    with transaction.atomic():
        for category in event_queryset.filter(parent_id=questionary.id, general_status='category'):
            category_questions = []
            for question in event_queryset.filter(parent_id=category.id, general_status='question'):
                supplier_answer = SupplierAnswer.objects.select_related('supplier', 'question', 'supplier_result', 'checker').filter(
                    question_id=question.id, supplier_id=supplier_id
                ).first()
                question_obj = {
                    'id': question.id,
                    'text': question.text,
                    'answer': question.answer,
                    'yes_no': question.yes_no,
                    'weight': question.weight,
                    "supplier_answer": {
                        'id': supplier_answer.id,
                        'answer': supplier_answer.answer,
                        'yes_no': supplier_answer.yes_no,
                        "weight": supplier_answer.weight
                    }
                }
                category_questions.append(question_obj)
            category_obj = {
                'id': category.id,
                'title': category.title,
                'weight': category.weight,
                'questions': category_questions
            }
            categories.append(category_obj)
    return {
        "is_submitted": supplier_result.is_submitted,
        "supplier_id": supplier_id,
        'supplier_name': supplier.name,
        "get_questionary_data": {
            "id": questionary.id,
            "title": questionary.title,
            "text": questionary.text,
            "success_weight": questionary.success_weight,
            "categories": categories
        }
    }


class SourcingRequestCategoryManager(APIView):
    permission_classes = (
        permissions.IsAuthenticated, IsCategoryManager | IsSourcingAdministrator | IsContractAdministrator |
        IsSourcingDirector
    )

    def get_sourcing_request_queryset(self):
        return SourcingRequest.objects.select_related(
            'organization', 'departement', 'categoryRequest', 'requestor', 'assigned_to', 'currency'
        )

    def post(self, request, *args, **kwargs):
        try:
            user = self.request.user
            request_data = self.request.data
            method = request_data.get('method')
            match method:
                case 'sourcing.request.list':
                    return Response(
                        sourcing_request_list_category_manager(request, user), status=status.HTTP_200_OK
                    )
                case 'supplier.answers':
                    return Response(
                        get_supplier_answers(request), status=status.HTTP_200_OK
                    )

        except Exception as e:
            return Response(exception_response(e), status=status.HTTP_400_BAD_REQUEST)



