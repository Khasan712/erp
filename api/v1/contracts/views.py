import os

import openpyxl
from rest_framework.response import Response
from openpyxl import Workbook
from django.db.models import Q, Sum, Avg, Count
from rest_framework.views import APIView
import pandas as pd
from rest_framework import status, permissions, generics
from docx2pdf import convert
from django.db import transaction
from django.core.exceptions import ValidationError
from api.v1.commons.pagination import make_pagination
from api.v1.commons.views import exception_response, get_serializer_errors, get_serializer_valid_response, not_serializer_is_valid, \
    object_not_found_response, serializer_valid_response
from api.v1.contracts.serializers import (
    ContractFileUploadSerializer,
    ContractSerializer,
    MassUploadSerializer,
    CategorySerializer,
    CurrencySerializer,
    ContractTypeSerializer,
    CostCenterSerializer,
    DepartementSerializer, DocumentContactSerializers, ContractTaskSerializers, ContractServiceSerializers,
    ContractCommoditySerializers, ContractConsultantSerializers, ContractListSerializers, ContractGetSerializer,
)
from api.v1.contracts.models import (
    Contract,
    Departement,
    Contract_Type,
    Category,
    Cost_Center,
    Currency, ContractTask, ConnectContractWithTask,
    DocumentContact
)
from api.v1.users.services import make_errors
from api.v1.users.permissions import (
    IsAdmin,
    IsContractAdministrator,
    IsSourcingDirector, IsSupplier, IsCategoryManager
)


# API for contracts :
class ContractStatusStatisticsView(APIView):
    permission_classes = (permissions.IsAuthenticated, IsSourcingDirector)

    def get_queryset(self):
        queryset = Contract.objects.select_related(
            'parent_agreement', 'departement', 'category', 'currency', 'organization', 'create_by', 'supplier'
        ).filter(organization_id=self.request.user.organization.id)
        return queryset

    def get(self, request):
        try:
            return Response(
                {
                    "success": True,
                    "message": 'Successfully got contracts status statistics.',
                    "error": [],
                    "data": {
                        'total':self.get_queryset().count(),
                        'active': self.get_queryset().filter(status='ACTIVE').count(),
                        'draft': self.get_queryset().filter(status='DRAFT').count(),
                        'expired': self.get_queryset().filter(status='EXPIRED').count()
                    },
                }, status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": 'Error occurred.',
                    "error": str(e),
                    "data": [],
                }, status=status.HTTP_400_BAD_REQUEST
            )


class ContractCategoryStatisticsView(APIView):
    permission_classes = (permissions.IsAuthenticated, IsSourcingDirector)

    def get_categories(self):
        categories = Category.objects.select_related('organization').filter(organization_id=self.request.user.organization.id)
        return categories

    def get(self, request):
        try:
            categories = self.get_categories().annotate(spends=Sum('contract__contract_amount')).values('name', 'spends')
            return Response(
                {
                    "success": True,
                    "message": 'Successfully got category statistics.',
                    "error": [],
                    "data": categories
                }, status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": 'Error occurred.',
                    "error": str(e),
                    "data": [],
                }, status=status.HTTP_400_BAD_REQUEST
            )


class ContractListView(APIView):
    permission_classes = (permissions.IsAuthenticated, IsContractAdministrator | IsSourcingDirector | IsCategoryManager)

    def get_queryset(self):
        queryset = Contract.objects.select_related('parent_agreement', 'departement', 'category', 'currency',
            'organization', 'create_by', 'supplier').filter(organization_id=self.request.user.organization.id)
        params = self.request.query_params
        contract_structure = params.get('contract_structure')
        category = params.get('category')
        currency = params.get('currency')
        terms = params.get('terms')
        contract_status = params.get('status')
        q = params.get('q')
        role = self.request.user.role
        if role == 'category_manager':
            queryset = queryset.filter(category_manager_id=self.request.user.id)
        if contract_structure:
            queryset = queryset.filter(contract_structure__in=contract_structure.split(','))
        if category:
            queryset = queryset.filter(category_id__in=category.split(','))
        if currency:
            queryset = queryset.filter(currency_id__in=currency.split(','))
        if terms:
            queryset = queryset.filter(terms__in=terms.split(','))
        if contract_status:
            queryset = queryset.filter(status__in=contract_status.split(','))
        if q:
            queryset = queryset.filter(
                Q(duration__icontains=q) | Q(name__icontains=q) | Q(contract_number__icontains=q) |
                Q(description__icontains=q) | Q(contract_structure__icontains=q) | Q(contract_amount__icontains=q) |
                Q(category__name__icontains=q) | Q(currency__name__icontains=q) | Q(terms__icontains=q) |
                Q(status__icontains=q) | Q(supplier__name__icontains=q)
            )
        return queryset

    def create_service_choices(self, service_choice, items, contract_id):
        creator = self.request.user.id
        with transaction.atomic():
            match service_choice:
                case 'service':
                    for service in items:
                        service['service'] = service.get('id')
                        service_serializer = ContractServiceSerializers(data=service)
                        if not service_serializer.is_valid():
                            get_serializer_errors(service_serializer)
                        service_serializer.save(contract_id=contract_id, creator_id=creator)
                case 'commodity':
                    for commodity in items:
                        commodity['commodity'] = commodity.get('id')
                        commodity_serializer = ContractCommoditySerializers(data=commodity)
                        if not commodity_serializer.is_valid():
                            get_serializer_errors(commodity_serializer)
                        commodity_serializer.save(contract_id=contract_id, creator_id=creator)
                case 'consultant':
                    for consultant in items:
                        consultant['consultant'] = consultant.get('id')
                        consultant_serializer = ContractConsultantSerializers(data=consultant)
                        if not consultant_serializer.is_valid():
                            get_serializer_errors(consultant_serializer)
                        consultant_serializer.save(contract_id=contract_id, creator_id=creator)

    def validate_contract_structure(self, data):
        contract_structure = data.get('contract_structure')
        match contract_structure:
            case 'Stand Alone':
                if data.get('parent_agreement') is not None:
                    return 'Can not be parent agreement.'
            case 'Master Agreement':
                if data.get('parent_agreement') is not None:
                    return 'Can not be parent agreement.'
            case 'Sub Agreement':
                if data.get('parent_agreement') is None:
                    return 'Choose parent agreement.'
                if data.get('parent_agreement') is not None:
                    contract_structure_of_parent_agreement = self.get_queryset().get(id=data.get('parent_agreement'))
                    if contract_structure_of_parent_agreement.contract_structure != 'Master Agreement':
                        return 'Contract structure must be Master Agreement in Parent agreement.'
            case '':
                return 'Choose contract structure.'
        return True

    def get(self, request):
        try:
            return Response(
                make_pagination(self.request, ContractListSerializers, self.get_queryset().order_by('-id')), status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response(
                exception_response(e), status=status.HTTP_400_BAD_REQUEST
            )

    def post(self, request):
        try:
            if self.request.user.role != 'contract_administrator':
                return Response(
                    {
                        'success': False,
                        'message': 'You do not have permission.',
                        'error': 'Error occurred.',
                        'data': []
                    }, status=status.HTTP_400_BAD_REQUEST
                )
            data = self.request.data
            service_choice = data.get('serviceCommodityConsultant')
            items = data.get('items')
            with transaction.atomic():
                if self.validate_contract_structure(data) is not True:
                    raise ValidationError(self.validate_contract_structure(data))
                serializer = ContractSerializer(data=data)
                if not serializer.is_valid():
                    raise ValidationError(message=f'{make_errors(serializer.errors)}')
                serializer.save(create_by_id=self.request.user.id, organization_id=self.request.user.organization.id)
                if data.get('documents'):
                    for document in data.get('documents'):
                        document['contract'] = serializer.data.get('id')
                        doc_serializer = DocumentContactSerializers(data=document)
                        if not doc_serializer.is_valid():
                            raise ValidationError(message=f'{make_errors(doc_serializer.errors)}')
                        doc_serializer.save()
                if not items and service_choice:
                    raise ValidationError(message="Please select any choices!")
                if service_choice and items:
                    self.create_service_choices(service_choice, items, serializer.data.get('id'))
        except Exception as e:
            return Response(
                exception_response(e), status=status.HTTP_400_BAD_REQUEST
            )
        else:
            return Response(
                serializer_valid_response(serializer), status=status.HTTP_201_CREATED
            )


class ContractMasterAgreementListView(APIView):
    permission_classes = (permissions.IsAuthenticated, IsContractAdministrator | IsSourcingDirector)

    def get_queryset(self):
        queryset = Contract.objects.select_related(
            'parent_agreement', 'departement', 'category', 'currency', 'organization', 'create_by', 'supplier'
        ).filter(organization_id=self.request.user.organization.id, contract_structure='Master Agreement')
        return queryset

    def get(self, request):
        try:
            return Response(
                make_pagination(self.request, ContractListSerializers, self.get_queryset()), status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response(exception_response(e), status=status.HTTP_400_BAD_REQUEST)


class ContractDetailView(APIView):
    permission_classes = (permissions.IsAuthenticated, IsSourcingDirector | IsContractAdministrator | IsSupplier | IsCategoryManager)

    def get_queryset(self):
        queryset = Contract.objects.select_related('parent_agreement', 'departement', 'category', 'currency',
            'organization', 'create_by', 'supplier').filter(organization_id=self.request.user.organization.id)
        return queryset

    def get_object(self):
        contract_id = self.request.query_params.get('contract')
        return self.get_queryset().filter(id=contract_id).first()

    def update_contract_tasks(self, tasks: list, contract_id: int):
        contract_tasks = ConnectContractWithTask.objects.select_related('contract', 'task', 'executor', 'checker').filter(
            contract_id=contract_id
        )
        with transaction.atomic():
            for task in tasks:
                contract_task = contract_tasks.filter(id=task.get('id')).first()
                contract_task.is_done = task.get('is_done')
                contract_task.save()

    def validate_contract(self):
        data = self.request.data
        contract_status = data.get('status')
        if self.request.user.role != 'contract_administrator':
            return 'You do not have permission!!!'
        if contract_status:
            if contract_status not in ('DRAFT', 'ACTIVE'):
                return f'Contract can not change to {contract_status}'
        if self.get_object() is None:
            return 'Contract not found.'
        return True
    
    def update_documents(self, documents: list, contract_id: int):
        contract = self.get_queryset().get(id=contract_id)
        with transaction.atomic():
            for document in documents:
                contract_document = DocumentContact.objects.select_related('contract').filter(id=document.get('id')).first()
                if contract_document:
                    document['contract'] = contract
                    serializer = DocumentContactSerializers(contract_document, data=document, partial=True)
                    if not serializer.is_valid():
                        raise ValidationError(not_serializer_is_valid(serializer.errors))
                    serializer.save()
                else:
                    pass

    def get(self, request):
        try:
            if self.get_object() is None:
                return Response(object_not_found_response())
            serializer = ContractGetSerializer(self.get_object())
            return Response(get_serializer_valid_response(serializer))
        except Exception as e:
            return Response(
                exception_response(e), status=status.HTTP_400_BAD_REQUEST
            )

    def patch(self, request):
        try:
            if not self.validate_contract():
                return Response(
                    {
                        "success": False,
                        "message": 'Error occurred.',
                        "error": self.validate_contract(),
                        "data": [],
                    }
                )
            data = self.request.data
            contract_tasks = data.get('tasks')
            documents = data.get('documents')
            service_commodity_consultant = data.get('serviceCommodityConsultant')
            items = data.get('items')
            serializer = ContractSerializer(self.get_object(), data=data, partial=True)
            with transaction.atomic():
                if not serializer.is_valid():
                    raise get_serializer_errors(serializer)
                serializer.save()
                if contract_tasks is not None:
                    self.update_contract_tasks(contract_tasks, serializer.data.get('id'))
                if documents:
                    self.update_documents(documents, serializer.data.get('id'))

        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": 'Error occurred.',
                    "error": str(e),
                    "data": [],
                }, status=status.HTTP_400_BAD_REQUEST
            )
        return Response(
            {
                "success": True,
                "message": 'Successfully updated Contract.',
                "error": [],
                "data": serializer.data,
            }, status=status.HTTP_200_OK
        )


class ProcessContractTaskView(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def post(self, request):
        try:
            with transaction.atomic():
                is_done = False
                contract = Contract.objects.select_related(
                    'parent_agreement', 'departement', 'category', 'currency', 'organization', 'create_by', 'supplier'
                )
                for task in self.request.data.get('tasks'):
                    task_id = ContractTask.objects.select_related('creator', 'contract', 'executor', 'checker').filter(
                        id=task['id'], contract__organization_id=self.request.user.organization.id
                    ).first()
                    contract = contract.filter(organization_id=self.request.user.orhanization.id, id=task.get('contract')).first()
                    if not contract:
                        raise ValidationError(message="Task not found.")
                    serializer = ContractTaskSerializers(task_id, data=task, partial=True)
                    if not serializer.is_valid():
                        raise ValidationError(message=f'{make_errors(serializer.errors)}')
                    is_done = task.get('is_done')
                    serializer.save(checker_id=self.request.user.id)
                # if is_done:
                #     contract.status = 'ACTIVE'
                #     contract.save()
        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": 'Error occurred.',
                    "error": str(e),
                    "data": [],
                }, status=status.HTTP_400_BAD_REQUEST
            )
        return Response(
            {
                "success": True,
                "message": 'Successfully done task events.',
                "error": [],
                "data": [],
            }, status=status.HTTP_201_CREATED
        )


class ContractFileUpload(generics.UpdateAPIView):
    queryset = Contract.objects.select_related('parent_agreement', 'departement', 'category', 'currency', 'organization', 'create_by', 'supplier')
    serializer_class = ContractFileUploadSerializer

    def patch(self, request, pk):
        item = Contract.objects.get(id=pk)
        item.document = request.data.get('document')
        item.save()
        item.document = convert(os.path.abspath(item.document.path))
        item.save()
        serializer = self.get_serializer(item)
        return Response(serializer.data)


class CurrencyListView(generics.ListCreateAPIView):
    queryset = Currency.objects.select_related('organization')
    serializer_class = CurrencySerializer

    def get(self, request):
        try:
            currencies = self.get_queryset().filter(organization_id=self.request.user.organization.id)
            serializer = self.get_serializer(currencies, many=True)
            return Response(
                {
                    "success": True,
                    "message": 'Service got successfully.',
                    "error": [],
                    "data": serializer.data,
                }, status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response(
                exception_response(e), status=status.HTTP_400_BAD_REQUEST
            )

    def post(self, request):

        serializer = CurrencySerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(organization = request.user.organization)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class CurrencyListDetailView(APIView):

    def get(self, request, id):
        queryset = Currency.objects.filter(organization = request.user).get(id=id)
        serializer = CurrencySerializer(queryset)
        return Response(serializer.data, status=status.HTTP_202_ACCEPTED)
    
    def put(self, request, id):
        queryset = Currency.objects.filter(organization = request.user).get(id=id)
        serializer = CurrencySerializer(queryset, data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, id):
        queryset = Currency.objects.filter(organization = request.user).get(id=id)
        queryset.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# API for cost center

class CostCenterListView(APIView):
    permission_classes = (permissions.IsAuthenticated,)
    def get(self, request):
        queryset = Cost_Center.objects.filter(organization = request.user)
        serializer = CostCenterSerializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_202_ACCEPTED)

    def post(self, request):

        serializer = CostCenterSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(organization = request.user.organization)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class CostCenterListDetailView(APIView):

    def get(self, request, id):
        queryset = Cost_Center.objects.filter(organization = request.user).get(id=id)
        serializer = CostCenterSerializer(queryset)
        return Response(serializer.data, status=status.HTTP_202_ACCEPTED)
    
    def put(self, request, id):
        queryset = Cost_Center.objects.filter(organization = request.user).get(id=id)
        serializer = CostCenterSerializer(queryset, data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, id):
        queryset = Cost_Center.objects.filter(organization = request.user).get(id=id)
        queryset.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


#API for contract type :

class ContractTypeListView(APIView):

    def get(self, request):
        queryset = Contract_Type.objects.filter(organization = request.user)
        serializer = ContractTypeSerializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_202_ACCEPTED)

    def post(self, request):

        serializer = ContractTypeSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(organization = request.user.organization)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ContractTypeListDetailView(APIView):

    def get(self, request, id):
        queryset = Contract_Type.objects.filter(organization = request.user).get(id=id)
        serializer = ContractTypeSerializer(queryset)
        return Response(serializer.data, status=status.HTTP_202_ACCEPTED)
    
    def put(self, request, id):
        queryset = Contract_Type.objects.filter(organization = request.user).get(id=id)
        serializer = ContractTypeSerializer(queryset, data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, id):
        queryset = Contract_Type.objects.filter(organization = request.user).get(id=id)
        queryset.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# API for department :
class DepartmentListView(generics.ListCreateAPIView):
    queryset = Departement.objects.select_related('organization', 'creator')
    serializer_class = DepartementSerializer

    def get(self, request):
        try:
            departments = self.get_queryset().filter(organization_id=self.request.user.organization.id)
            serializer = self.get_serializer(departments, many=True)
            return Response(
                {
                    "success": True,
                    "message": 'Department got successfully.',
                    "error": [],
                    "data": serializer.data,
                }, status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response(
                exception_response(e), status=status.HTTP_400_BAD_REQUEST
            )

    def post(self, request):

        serializer = DepartementSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(organization = request.user.organization)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class DepartmentListDetailView(APIView):

    def get(self, request, id):
        queryset = Departement.objects.filter(organization = request.user).get(id=id)
        serializer = DepartementSerializer(queryset)
        return Response(serializer.data, status=status.HTTP_202_ACCEPTED)
    
    def put(self, request, id):
        queryset = Departement.objects.filter(organization = request.user).get(id=id)
        serializer = DepartementSerializer(queryset, data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, id):
        queryset = Departement.objects.filter(organization = request.user).get(id=id)
        queryset.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# APi for category :
class CategoryListView(APIView):
    permission_classes = (permissions.IsAuthenticated, IsSourcingDirector | IsContractAdministrator | IsSupplier | IsCategoryManager)

    def get_queryset(self):
        queryset = Category.objects.select_related('organization').filter(
            organization_id=self.request.user.organization.id)
        return queryset

    def get(self, request):
        try:
            return Response(make_pagination(self.request, CategorySerializer, self.get_queryset()))
        except Exception as e:
            return Response(exception_response(e), status=status.HTTP_400_BAD_REQUEST)

    def post(self, request):

        serializer = CategorySerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(organization = request.user.organization)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class CategoryListDetailView(APIView):

    def get(self, request, id):
        queryset = Category.objects.filter(organization = request.user).get(id=id)
        serializer = CategorySerializer(queryset)
        return Response(serializer.data, status=status.HTTP_202_ACCEPTED)
    
    def put(self, request, id):
        queryset = Category.objects.filter(organization = request.user).get(id=id)
        serializer = CategorySerializer(queryset, data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, id):
        queryset = Category.objects.filter(organization = request.user).get(id=id)
        queryset.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class ContractListDetailView(APIView):

    def get(self, request, id):
        con = Contract.objects.filter(organization = request.user).get(id=id)
        serializer = ContractSerializer(con)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def put(self, request, id):
        con = Contract.objects.filter(organization = request.user).get(id=id)
        serializer = ContractSerializer(con, data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_202_ACCEPTED)
        else:
            return Response(serializer.errors, status = status.HTTP_400_BAD_REQUEST)

    def delete(self, request, id):
        con = Contract.objects.filter(organization = request.user).get(id=id)
        con.delete()
        return Response(status = status.HTTP_204_NO_CONTENT)


class MassUploadView(APIView):

    def post(self, request):
        serializer = MassUploadSerializer(data=request.data)
        file = request.data['file']
        file = pd.read_excel(file)
        file = pd.DataFrame(data=file)
        for i in range(0, len(file)):
            Contract.objects.create(name=file.iloc[i][0], description=file.iloc[i][1], organization=request.user.organization)
        return Response("ok")


from openpyxl.styles import Alignment, Font, Protection


class ContractToExelApi(APIView):
    permission_classes = (permissions.IsAuthenticated, )

    def get_queryset(self):
        queryset = Contract.objects.select_related(
            'parent_agreement', 'departement', 'category', 'currency', 'organization', 'create_by', 'supplier'
        ).filter(
            organization_id=self.request.user.organization.id
        )
        return queryset

    def get_contract_fields(self):
        fields = {
            'category_manager': 'Category manager',
            'contract_owner': 'Contract owner',
            'lawyer': 'Lawyer',
            'project_owner': 'Project owner',
            'creation_date': 'Creation date',
            'effective_date': 'Effective date',
            'expiration_date': 'Expiration date',
            'duration': 'Duration',
            'name': 'Name',
            'contract_number': 'Contract number',
            'description': 'Description',
            'parent_agreement': 'Parent agreement',
            'departement': 'Department',
            'contract_structure': 'Contract structure',
            'contract_amount': 'Contract amount',
            'category': 'Category',
            'currency': 'Currency',
            'organization': 'Organization',
            'create_by': 'Create by',
            'terms': 'Terms',
            'contract_notice': 'Contract notice',
            'amendment': 'Amendment',
            'status': 'Status',
            'count_changes': 'Count changes',
            'notification': 'Notification',
            'supplier': 'Supplier',
        }
        return fields

    def get_filtered_data(self):
        data = self.request.data
        fields = data.get('fields')
        excel_header = []
        if not fields:
            return 'Choose one of the field.'
        for field in fields:
            if field not in self.get_contract_fields().keys():
                return 'Choose valid field.'
            excel_header.append(self.get_contract_fields()[field])
        contract_data = {
            'excel_header': excel_header,
            'queryset_fields': fields,
            'queryset': self.get_queryset()
        }
        return contract_data

    def get_rows(self, contract):
        row = []
        for field in self.get_filtered_data()['queryset_fields']:
            match field:
                case 'category_manager':
                    row.append(contract.category_manager.first_name)
                case 'contract_owner':
                    row.append(contract.contract_owner.first_name)
                case 'lawyer':
                    row.append(contract.lawyer.first_name)
                case 'project_owner':
                    row.append(contract.project_owner.first_name)
                case 'creation_date':
                    row.append(contract.creation_date.strftime('%m/%d/%Y, %H:%M'))
                case 'effective_date':
                    row.append(contract.effective_date.strftime('%m/%d/%Y'))
                case 'expiration_date':
                    row.append(contract.expiration_date.strftime('%m/%d/%Y'))
                case 'duration':
                    row.append(contract.duration)
                case 'name':
                    row.append(contract.name)
                case 'contract_number':
                    row.append(contract.contract_number)
                case 'description':
                    row.append(contract.description)
                case 'parent_agreement':
                    row.append(contract.parent_agreement.name)
                case 'departement':
                    row.append(contract.departement.name)
                case 'contract_structure':
                    row.append(contract.contract_structure)
                case 'contract_amount':
                    row.append(contract.contract_amount)
                case 'category':
                    row.append(contract.category.name)
                case 'currency':
                    row.append(contract.currency.name)
                case 'create_by':
                    row.append(contract.create_by.first_name)
                case 'terms':
                    row.append(contract.terms)
                case 'contract_notice':
                    row.append(contract.contract_notice)
                case 'amendment':
                    row.append(contract.amendment)
                case 'status':
                    row.append(contract.status)
                case 'count_changes':
                    row.append(contract.count_changes)
                case 'notification':
                    row.append(contract.notification)
                case 'supplier':
                    row.append(contract.supplier.name)
        return row

    def post(self, request):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.sheet_properties.tabColor = '1072BA'
        worksheet.freeze_panes = 'I2'

        contract_queryset = self.get_filtered_data()['queryset']
        columns = self.get_filtered_data()['excel_header']
        row_num = 1
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(bold=True)
        for _, contract in enumerate(contract_queryset, 1):
            row_num += 1
            row = self.get_rows(contract)
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.protection = Protection(locked=True)
        workbook.save('file.xlsx')
        return Response(
            {
                'file': 'f',
            }
        )


class ContractImportDataFromExelApi(APIView):
    permission_classes = (permissions.IsAuthenticated, )

    def get_queryset(self):
        queryset = Contract.objects.select_related(
            'parent_agreement', 'departement', 'category', 'currency', 'organization', 'create_by', 'supplier'
        ).filter(
            organization_id=self.request.user.organization.id
        )
        return queryset

    def get_contract_fields(self):
        fields = {
            'category_manager': 'Category manager',
            'contract_owner': 'Contract owner',
            'lawyer': 'Lawyer',
            'project_owner': 'Project owner',
            'creation_date': 'Creation date',
            'effective_date': 'Effective date',
            'expiration_date': 'Expiration date',
            'duration': 'Duration',
            'name': 'Name',
            'contract_number': 'Contract number',
            'description': 'Description',
            'parent_agreement': 'Parent agreement',
            'departement': 'Department',
            'contract_structure': 'Contract structure',
            'contract_amount': 'Contract amount',
            'category': 'Category',
            'currency': 'Currency',
            'organization': 'Organization',
            'create_by': 'Create by',
            'terms': 'Terms',
            'contract_notice': 'Contract notice',
            'amendment': 'Amendment',
            'status': 'Status',
            'count_changes': 'Count changes',
            'notification': 'Notification',
            'supplier': 'Supplier',
        }
        return fields

    def post(self, request):
        file = request.data.get('file')
        if not file:
            return Response('Upload file.')
        wb_obj = openpyxl.load_workbook(file)
        sheet_obj = wb_obj.active
        max_col = sheet_obj.max_column
        queryset_fields = []
        key_list = list(self.get_contract_fields().keys())
        value_list = list(self.get_contract_fields().values())
        for i in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=1, column=i)
            if cell_obj.value not in self.get_contract_fields().values():
                print(f'Rename {cell_obj.value}.')
            field_name = value_list.index(cell_obj.value)
            queryset_fields.append(key_list[field_name])

        for i in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=2, column=i)
        return Response(
            {
                'data': sheet_obj.max_column
            }
        )




