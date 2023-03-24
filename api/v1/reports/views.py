from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Protection
from rest_framework import permissions
from rest_framework.views import APIView
from rest_framework.response import Response

from .report_fields import REPORT_FIELDS
from api.v1.contracts.models import Contract
from .models import Report
import os

from ..services.models import Service


class ReportAPi(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def get_contract_queryset(self):
        return Contract.objects.select_related(
            'parent_agreement', 'departement', 'category', 'currency', 'organization', 'create_by', 'supplier'
        ).filter(
            organization_id=self.request.user.organization.id
        )

    def contract_to_exel(self):
        request_body = self.request.data
        request_fields = request_body.get("fields")
        if not request_fields:
            raise ValueError("Exel header not given!")
        exel_headers = dict()
        contract_fields = list(dict.fromkeys(request_fields.get("contract")))
        for contract_field in contract_fields:

            if contract_field not in REPORT_FIELDS['contract'].keys():
                raise ValueError(f"Given contract variable not found! `{contract_field}`")

            if contract_field == 'category_manager':
                category_manager_fields = request_fields.get('category_manager')
                if not category_manager_fields:
                    raise ValueError("Category manager header not given!")
                category_manager_fields = list(dict.fromkeys(category_manager_fields))
                collect_category_manager_fields = dict()
                for category_manager_field in category_manager_fields:
                    if category_manager_field not in REPORT_FIELDS['user'].keys():
                        raise ValueError("Category manager variable not found!")
                    collect_category_manager_fields[category_manager_field] = f"Category manager {REPORT_FIELDS['user'][category_manager_field]}"
                exel_headers['category_manager'] = collect_category_manager_fields

            elif contract_field == 'serviceCommodityConsultant':
                service_fields = request_fields.get("service")
                if not service_fields:
                    raise ValueError("Service header not given!")
                service_fields = list(dict.fromkeys(service_fields))
                collect_service_fields = dict()
                for service_field in service_fields:
                    if service_field not in REPORT_FIELDS['service'].keys():
                        raise ValueError("Service variable not found!")
                    collect_service_fields[service_field] = f"Service {REPORT_FIELDS['service'][service_field]}"
                exel_headers['service'] = collect_service_fields
            else:
                exel_headers[contract_field] = REPORT_FIELDS['contract'][contract_field]

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.sheet_properties.tabColor = '1072BA'
        worksheet.freeze_panes = 'I2'

        contract_queryset = self.get_contract_queryset()

        row_num = 1
        col_num = 1
        for column_key, column_val in exel_headers.items():
            if column_key == 'category_manager':
                for c_m_title in column_val.values():
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = c_m_title
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.font = Font(bold=True)
                    col_num+=1
            elif column_key == 'service':
                for s_title in column_val.values():
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = s_title
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.font = Font(bold=True)
                    col_num+=1
            else:
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_val
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(bold=True)
                col_num+=1

        col_num = 1
        max_down_row_num = 2
        row_num = max_down_row_num
        for c in range(len(contract_queryset)):
            for column_key, column_val in exel_headers.items():
                if column_key == 'category_manager':
                    val = contract_queryset[c].category_manager.__dict__
                    for c_m_key in column_val.keys():
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = val[c_m_key]
                        cell.protection = Protection(locked=True)
                        col_num+=1

                elif column_key == 'service':
                    all_services = Service.objects.select_related("organization", 'creator').filter(
                        contract_services__contract_id=contract_queryset[c].id
                    )
                    s_row = row_num
                    s_col = col_num
                    for all_s in range(len(all_services)):
                        s_obj_dict = all_services[all_s].__dict__
                        for s_key in column_val.keys():
                            cell = worksheet.cell(row=s_row, column=s_col)
                            cell.value = s_obj_dict[s_key]
                            cell.protection = Protection(locked=True)
                            s_col+=1
                        s_row+=1
                        s_col = col_num
                        max_down_row_num+=1
                    col_num += len(column_val.keys())
                else:
                    c_val = contract_queryset[c].__dict__[column_key]
                    if column_key == 'creation_date':
                        c_val = contract_queryset[c].creation_date.strftime('%m/%d/%Y, %H:%M')
                    elif column_key == 'effective_date':
                        c_val = contract_queryset[c].effective_date.strftime('%m/%d/%Y')
                    elif column_key == 'expiration_date':
                        c_val = contract_queryset[c].expiration_date.strftime('%m/%d/%Y')
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = c_val
                    cell.protection = Protection(locked=True)
                    col_num+=1
            max_down_row_num += 1
            row_num = max_down_row_num
            col_num = 1
            print(row_num)

        print(os.path.join(workbook.save('file.xlsx')))
        # file_save = Report.objects.create(
        #     done_by_id=self.request.user.id,
        #     report_file=workbook.save('file.xlsx'),
        # )
        # return file_save.report_file

    def post(self, request):
        try:
            request_data = request.data
            method = request_data.get('method')
            if not method or method not in ('contract.to.exel',):
                raise ValueError("Method not given or not found!")
            match method:
                case 'contract.to.exel':
                    """ Write contract data to exel """
                    report_file = self.contract_to_exel()
                    return Response({
                        "success": True,
                        # "file": report_file
                    })
        except Exception as e:
            return Response({
                "success": False,
                "error": str(e)
            })

