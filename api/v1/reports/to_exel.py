import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Protection

from .models import Report
from .report_fields import REPORT_FIELDS
from ..contracts.models import Contract
from ..services.models import Service
import os

def get_contract_queryset(request):
    return Contract.objects.select_related(
        'parent_agreement', 'departement', 'category', 'currency', 'organization', 'create_by', 'supplier'
    ).filter(
        organization_id=request.user.organization.id
    )


def contract_to_exel(request):
    user = request.user
    request_body = request.data
    request_fields = request_body.get("fields")
    if not request_fields:
        raise ValueError("Exel header not given!")
    exel_headers = {}
    for contract_field in dict.fromkeys(request_fields.get('contract', [])):

        if contract_field not in REPORT_FIELDS['contract']:
            raise ValueError(f"Given contract variable not found! `{contract_field}`")

        if contract_field in ['category_manager', 'contract_owner', 'lawyer', 'project_owner']:
            contract_user_fields = request_fields.get(contract_field)
            if not contract_user_fields:
                raise ValueError(f"{contract_field} header not given!")
            contract_user_manager_fields = {}
            for contract_user_field in dict.fromkeys(contract_user_fields):
                if contract_user_field not in REPORT_FIELDS['user']:
                    raise ValueError(f"{contract_field} variable not found!")
                contract_user_manager_fields[contract_user_field] = f"{contract_field.replace('_',' ').capitalize()} {REPORT_FIELDS['user'][contract_user_field]}"
            exel_headers[contract_field] = contract_user_manager_fields

        elif contract_field in ['departement', 'category', 'currency', 'supplier']:
            contract_d_c_c_s_fields = request_fields.get(contract_field)
            if not contract_d_c_c_s_fields:
                raise ValueError(f"{contract_field} header not given!")
            contract_d_c_c_s_manager_fields = {}
            for contract_d_c_c_s_field in dict.fromkeys(contract_d_c_c_s_fields):
                if contract_d_c_c_s_field not in REPORT_FIELDS[contract_field]:
                    raise ValueError(f"{contract_field} variable not found!")
                contract_d_c_c_s_manager_fields[
                    contract_d_c_c_s_field] = f"{contract_field.replace('_', ' ').capitalize()} {REPORT_FIELDS[contract_field][contract_d_c_c_s_field]}"
            exel_headers[contract_field] = contract_d_c_c_s_manager_fields

        elif contract_field == 'serviceCommodityConsultant':
            service_fields = request_fields.get("service")
            commodity_fields = request_fields.get("service")
            consultant_fields = request_fields.get("service")
            if not service_fields:
                raise ValueError("Service header not given!")
            collect_service_fields = {}
            for service_field in dict.fromkeys(service_fields):
                if service_field not in REPORT_FIELDS['service'].keys():
                    raise ValueError("Service variable not found!")
                collect_service_fields[service_field] = f"Service {REPORT_FIELDS['service'][service_field]}"
            exel_headers['service'] = collect_service_fields
        else:
            exel_headers[contract_field] = REPORT_FIELDS['contract'][contract_field]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.sheet_properties.tabColor = '1072BA'
    worksheet.freeze_panes = 'I2'

    contract_queryset = get_contract_queryset(request)

    row_num = 1
    col_num = 1
    for column_key, column_val in exel_headers.items():
        if column_key in ['category_manager', 'contract_owner', 'service']:
            for title in column_val.values():
                cell = worksheet.cell(row=row_num, column=col_num, value=title)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(bold=True)
                col_num+=1
        else:
            cell = worksheet.cell(row=row_num, column=col_num, value=column_val)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(bold=True)
            col_num+=1

    col_num = 1
    max_down_row_num = 2
    row_num = max_down_row_num
    for c, q_v in enumerate(contract_queryset):
        print(c)
        for column_key, column_val in exel_headers.items():
            if column_key in ['category_manager', 'contract_owner']:
                val = contract_queryset[c].category_manager.__dict__
                for c_m_key in column_val.keys():
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = val[c_m_key]
                    cell.protection = Protection(locked=True)
                    col_num+=1
            elif column_key == 'service':
                all_services = Service.objects.select_related("organization", 'creator').filter(
                    contract_services__contract_id=contract_queryset[c].id
                )
                s_row = row_num
                s_col = col_num
                for q_i, q_v in enumerate(all_services):
                    s_obj_dict = all_services[q_i].__dict__
                    for s_key in column_val.keys():
                        cell = worksheet.cell(row=s_row, column=s_col)
                        cell.value = s_obj_dict[s_key]
                        cell.protection = Protection(locked=True)
                        s_col+=1
                    s_row+=1
                    s_col = col_num
                    max_down_row_num+=1
                col_num += len(column_val.keys())
            else:
                c_val = contract_queryset[c].__dict__[column_key]
                if column_key == 'creation_date':
                    c_val = contract_queryset[c].creation_date.strftime('%m/%d/%Y, %H:%M')
                elif column_key == 'effective_date':
                    c_val = contract_queryset[c].effective_date.strftime('%m/%d/%Y')
                elif column_key == 'expiration_date':
                    c_val = contract_queryset[c].expiration_date.strftime('%m/%d/%Y')
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = c_val
                cell.protection = Protection(locked=True)
                col_num+=1
        max_down_row_num += 1
        row_num = max_down_row_num
        col_num = 1

    current_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S:%f')
    new_path = r'media/reports/%s/to_exel' % user.organization.name
    if not os.path.exists(new_path):
        os.makedirs(new_path)
    workbook.save(f'{new_path}/{current_time}.xlsx')
    file_save = Report.objects.create(
        report_model='contract',
        done_by_id=user.id,
        report_file=f'{new_path}/{current_time}.xlsx',
    )
    return f'{file_save.report_file}'
